

import argparse
import re
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ════════════════════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════════════════════

RATING_COL_CANDIDATES   = ['star rating', 'rating', 'stars', 'score']
TEXT_COL_CANDIDATES     = ['review content', 'review', 'comment', 'feedback', 'text']
DATE_COL_CANDIDATES     = ['review date', 'date', 'created', 'timestamp']
GUIDE_COL_CANDIDATES    = ['local guide', 'guide', 'verified']
AUTHOR_REVIEWS_CANDIDATES = ['author reviews', 'reviewer reviews', 'total reviews']

NEGATIVE_THRESHOLD = 2   # ratings <= this are "negative"
POSITIVE_THRESHOLD = 4   # ratings >= this are "positive"

# Reused keyword categories for negative-review theme breakdown
CATEGORIES = {
    'Staff Behaviour / Rudeness':    r'rude|misbehav|arrogant|bad behav|behaviour|behav|language|insensitive|harass|abuse|shout|yell|disrespect',
    'Waiting Time / Delays':         r'wait|delay|slow|time|hours|whole day|queue|long|late|come after|not on time|no response',
    'Management / Administration':   r'manage|administ|coordinat|system|process|protocol|paperwork|file|sign|approval|counter|mismanage',
    'Staff Availability / Shortage': r'no staff|less staff|staff shortage|not available|absent|no one|nobody|no doctor|no nurse|PC boy|ward boy',
    'Communication / Information':   r'inform|explain|communicat|answer|respond|no guidance|guide|told|said|unclear|confusion|proper answer',
    'Hygiene / Cleanliness':         r'clean|dirty|washroom|bathroom|hygiene|toilet|sanit',
    'Nursing Care':                   r'nurse|sister|ward nurse|nursing|ICU nurse|nurses behav',
    'Billing / Costs':                r'cost|expensive|charge|money|bill|fee|costly|overcharge|loot|double',
    'Technology / Network':           r'network|internet|system down|server|technical|IT|connectivity|online',
    'Patient Empathy / Emotional':   r'empath|scare|frighten|sensitive|courage|humanity|compassion|mental|distress|trauma',
}

# ── Visual style ───────────────────────────────────────────────────────────
DARK_MAROON  = '6B1A1A'
MED_MAROON   = '8B2E2E'
ACCENT_GOLD  = 'C9A84C'
GREY_ROW     = 'F5F5F5'
RATING_COLORS = {1: 'C00000', 2: 'ED7D31', 3: 'FFC000', 4: '9BBB59', 5: '4F8B3B'}
CHART_PALETTE = ['#6B1A1A', '#C9A84C', '#4472C4', '#70AD47', '#ED7D31', '#8B2E2E', '#4F8B3B']

plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.edgecolor'] = '#888888'
plt.rcParams['axes.grid'] = True
plt.rcParams['grid.alpha'] = 0.25


# ════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ════════════════════════════════════════════════════════════════════════════

def find_column(df, candidates):
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        for col_lower, original in cols_lower.items():
            if cand in col_lower:
                return original
    return None


def load_data(input_path, rating_col=None, text_col=None, date_col=None):
    path = Path(input_path)
    df = pd.read_csv(path) if path.suffix.lower() == '.csv' else pd.read_excel(path)

    rating_col = rating_col or find_column(df, RATING_COL_CANDIDATES)
    text_col   = text_col or find_column(df, TEXT_COL_CANDIDATES)
    date_col   = date_col or find_column(df, DATE_COL_CANDIDATES)
    guide_col  = find_column(df, GUIDE_COL_CANDIDATES)
    author_reviews_col = find_column(df, AUTHOR_REVIEWS_CANDIDATES)

    if rating_col is None:
        raise ValueError(f"Could not find a rating column. Columns found: {list(df.columns)}")

    rename_map = {rating_col: '_rating'}
    if text_col:   rename_map[text_col] = '_text'
    if date_col:   rename_map[date_col] = '_date'
    if guide_col:  rename_map[guide_col] = '_local_guide'
    if author_reviews_col: rename_map[author_reviews_col] = '_author_reviews'
    df = df.rename(columns=rename_map)

    df['_rating'] = pd.to_numeric(df['_rating'], errors='coerce')
    df = df[df['_rating'].notna()].copy()
    df['_rating'] = df['_rating'].astype(int)

    if '_text' in df.columns:
        df['_text'] = df['_text'].fillna('').astype(str)
    else:
        df['_text'] = ''

    if '_date' in df.columns:
        df['_date'] = pd.to_datetime(df['_date'], errors='coerce')

    df = df.reset_index(drop=True)
    return df


def categorize(text):
    text_lower = text.lower()
    matched = [cat for cat, pattern in CATEGORIES.items() if re.search(pattern, text_lower)]
    return matched if matched else ['Other / General']


# ════════════════════════════════════════════════════════════════════════════
# STAT COMPUTATIONS
# ════════════════════════════════════════════════════════════════════════════

def compute_rating_distribution(df):
    dist = df['_rating'].value_counts().sort_index()
    full = pd.Series({i: dist.get(i, 0) for i in range(1, 6)})
    pct = (full / full.sum() * 100).round(1)
    return pd.DataFrame({'Star Rating': full.index, 'Count': full.values, 'Percent': pct.values})


def compute_summary_stats(df):
    n = len(df)
    avg = df['_rating'].mean()
    median = df['_rating'].median()
    neg = (df['_rating'] <= NEGATIVE_THRESHOLD).sum()
    pos = (df['_rating'] >= POSITIVE_THRESHOLD).sum()
    neutral = n - neg - pos
    stats = [
        ('Total Reviews', n),
        ('Average Rating', round(avg, 2)),
        ('Median Rating', median),
        (f'Negative Reviews (\u2264{NEGATIVE_THRESHOLD}\u2605)', f'{neg} ({neg/n*100:.1f}%)'),
        ('Neutral Reviews (3\u2605)', f'{neutral} ({neutral/n*100:.1f}%)'),
        (f'Positive Reviews (\u2265{POSITIVE_THRESHOLD}\u2605)', f'{pos} ({pos/n*100:.1f}%)'),
    ]
    if '_local_guide' in df.columns:
        guides = df['_local_guide'].sum()
        stats.append(('Local Guide Reviewers', f'{guides} ({guides/n*100:.1f}%)'))
    if '_date' in df.columns and df['_date'].notna().any():
        stats.append(('Date Range', f"{df['_date'].min().date()} to {df['_date'].max().date()}"))
    return stats


def compute_monthly_trend(df):
    if '_date' not in df.columns or df['_date'].isna().all():
        return None
    d = df.dropna(subset=['_date']).copy()
    d['_month'] = d['_date'].dt.to_period('M').astype(str)
    grp = d.groupby('_month')['_rating'].agg(['mean', 'count']).reset_index()
    grp.columns = ['Month', 'Average Rating', 'Review Count']
    grp['Average Rating'] = grp['Average Rating'].round(2)
    return grp.sort_values('Month')


def compute_theme_breakdown(df):
    neg = df[df['_rating'] <= NEGATIVE_THRESHOLD].copy()
    neg = neg[neg['_text'].str.strip() != '']
    if len(neg) == 0:
        return None
    neg['_categories'] = neg['_text'].apply(categorize)
    freq = {}
    for cats in neg['_categories']:
        for c in cats:
            freq[c] = freq.get(c, 0) + 1
    freq_df = pd.DataFrame(sorted(freq.items(), key=lambda x: -x[1]), columns=['Theme', 'Mentions'])
    freq_df['% of Negative Reviews'] = (freq_df['Mentions'] / len(neg) * 100).round(1)
    return freq_df, len(neg)


def compute_review_length_by_rating(df):
    d = df[df['_text'].str.strip() != ''].copy()
    if len(d) == 0:
        return None
    d['_len'] = d['_text'].str.split().str.len()
    grp = d.groupby('_rating')['_len'].agg(['mean', 'median', 'count']).reset_index()
    grp.columns = ['Star Rating', 'Avg Word Count', 'Median Word Count', 'Reviews with Text']
    grp['Avg Word Count'] = grp['Avg Word Count'].round(1)
    return grp


def compute_reviewer_profile(df):
    rows = []
    if '_local_guide' in df.columns:
        for is_guide in [True, False]:
            sub = df[df['_local_guide'] == is_guide]
            if len(sub) == 0:
                continue
            label = 'Local Guide' if is_guide else 'Regular Reviewer'
            rows.append((label, len(sub), round(sub['_rating'].mean(), 2)))
    if '_author_reviews' in df.columns:
        df['_reviewer_type'] = pd.cut(
            df['_author_reviews'].fillna(0),
            bins=[-1, 1, 10, 100, float('inf')],
            labels=['First-time (1)', 'Occasional (2-10)', 'Active (11-100)', 'Prolific (100+)']
        )
        for label, sub in df.groupby('_reviewer_type', observed=True):
            if len(sub) == 0:
                continue
            rows.append((str(label), len(sub), round(sub['_rating'].mean(), 2)))
    return rows


def compute_dow_pattern(df):
    if '_date' not in df.columns or df['_date'].isna().all():
        return None
    d = df.dropna(subset=['_date']).copy()
    d['_dow'] = d['_date'].dt.day_name()
    order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    grp = d.groupby('_dow')['_rating'].agg(['mean', 'count']).reindex(order).reset_index()
    grp.columns = ['Day of Week', 'Average Rating', 'Review Count']
    grp['Average Rating'] = grp['Average Rating'].round(2)
    return grp


# ════════════════════════════════════════════════════════════════════════════
# CHART GENERATION (matplotlib PNGs)
# ════════════════════════════════════════════════════════════════════════════

def save_chart(fig, path):
    fig.tight_layout()
    fig.savefig(path, dpi=150, facecolor='white')
    plt.close(fig)


def chart_rating_distribution(dist_df, out_path):
    fig, ax = plt.subplots(figsize=(6, 4))
    colors = [f"#{RATING_COLORS[s]}" for s in dist_df['Star Rating']]
    ax.bar(dist_df['Star Rating'].astype(str) + '\u2605', dist_df['Count'], color=colors)
    for i, v in enumerate(dist_df['Count']):
        ax.text(i, v + max(dist_df['Count']) * 0.01, str(v), ha='center', fontsize=9, fontweight='bold')
    ax.set_title('Rating Distribution', fontsize=13, fontweight='bold', color='#333')
    ax.set_ylabel('Number of Reviews')
    ax.spines[['top', 'right']].set_visible(False)
    save_chart(fig, out_path)


def chart_monthly_trend(trend_df, out_path):
    fig, ax1 = plt.subplots(figsize=(9, 4))
    ax2 = ax1.twinx()
    ax2.bar(trend_df['Month'], trend_df['Review Count'], color='#C9A84C', alpha=0.35, label='Review Count')
    ax1.plot(trend_df['Month'], trend_df['Average Rating'], color='#6B1A1A', marker='o', linewidth=2, label='Avg Rating')
    ax1.set_ylim(0.8, 5.2)
    ax1.set_ylabel('Average Rating', color='#6B1A1A')
    ax2.set_ylabel('Review Count', color='#C9A84C')
    ax1.set_title('Monthly Rating Trend', fontsize=13, fontweight='bold', color='#333')
    ax1.tick_params(axis='x', rotation=45)
    if len(trend_df) > 18:
        for label in ax1.xaxis.get_ticklabels()[::2]:
            label.set_visible(False)
    ax1.spines[['top']].set_visible(False)
    save_chart(fig, out_path)


def chart_theme_breakdown(freq_df, out_path):
    fig, ax = plt.subplots(figsize=(7, 4.5))
    freq_df = freq_df.sort_values('Mentions')
    ax.barh(freq_df['Theme'], freq_df['Mentions'], color='#6B1A1A')
    for i, v in enumerate(freq_df['Mentions']):
        ax.text(v + 0.3, i, str(v), va='center', fontsize=9)
    ax.set_title('Negative Review Themes', fontsize=13, fontweight='bold', color='#333')
    ax.set_xlabel('Mentions')
    ax.spines[['top', 'right']].set_visible(False)
    save_chart(fig, out_path)


def chart_review_length(len_df, out_path):
    fig, ax = plt.subplots(figsize=(6, 4))
    colors = [f"#{RATING_COLORS.get(int(s), '888888')}" for s in len_df['Star Rating']]
    ax.bar(len_df['Star Rating'].astype(str) + '\u2605', len_df['Avg Word Count'], color=colors)
    ax.set_title('Avg Review Length by Rating', fontsize=13, fontweight='bold', color='#333')
    ax.set_ylabel('Average Word Count')
    ax.spines[['top', 'right']].set_visible(False)
    save_chart(fig, out_path)


def chart_reviewer_profile(profile_rows, out_path):
    fig, ax = plt.subplots(figsize=(7, 4))
    labels = [r[0] for r in profile_rows]
    counts = [r[1] for r in profile_rows]
    avgs = [r[2] for r in profile_rows]
    x = range(len(labels))
    bars = ax.bar(x, counts, color=CHART_PALETTE[:len(labels)])
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=20, ha='right')
    ax.set_ylabel('Number of Reviews')
    for i, (c, a) in enumerate(zip(counts, avgs)):
        ax.text(i, c + max(counts) * 0.01, f'avg {a}\u2605', ha='center', fontsize=8)
    ax.set_title('Reviewer Profile Breakdown', fontsize=13, fontweight='bold', color='#333')
    ax.spines[['top', 'right']].set_visible(False)
    save_chart(fig, out_path)


def chart_dow_pattern(dow_df, out_path):
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.bar(dow_df['Day of Week'], dow_df['Average Rating'], color='#4472C4')
    ax.set_ylim(0, 5.5)
    ax.set_title('Average Rating by Day of Week', fontsize=13, fontweight='bold', color='#333')
    ax.spines[['top', 'right']].set_visible(False)
    save_chart(fig, out_path)


# ════════════════════════════════════════════════════════════════════════════
# EXCEL STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def header_font(sz=11):
    return Font(name='Arial', size=sz, bold=True, color='FFFFFF')


def body_font(sz=10, bold=False):
    return Font(name='Arial', size=sz, bold=bold)


def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)


def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)


def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def merge_title(ws, cell_range, text, bg=DARK_MAROON, sz=13):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(':')[0]]
    c.value = text
    c.font = Font(name='Arial', size=sz, bold=True, color='FFFFFF')
    c.fill = fill(bg)
    c.alignment = center()


def write_table(ws, df, start_row, start_col=1, header_bg=MED_MAROON):
    for j, col_name in enumerate(df.columns):
        c = ws.cell(row=start_row, column=start_col + j, value=col_name)
        c.font = header_font(10)
        c.fill = fill(header_bg)
        c.alignment = center()
        c.border = thin_border()
    for i, row in enumerate(df.itertuples(index=False)):
        r = start_row + 1 + i
        bg = GREY_ROW if i % 2 == 0 else 'FFFFFF'
        for j, val in enumerate(row):
            c = ws.cell(row=r, column=start_col + j, value=val)
            c.font = body_font()
            c.alignment = left() if j == 0 else center()
            c.fill = fill(bg)
            c.border = thin_border()
    return start_row + 1 + len(df)


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def embed_image(ws, img_path, anchor_cell):
    if Path(img_path).exists():
        img = XLImage(str(img_path))
        img.width = img.width * 0.62
        img.height = img.height * 0.62
        ws.add_image(img, anchor_cell)


# ════════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ════════════════════════════════════════════════════════════════════════════

def run_analysis(input_path, output_dir):
    output_dir = Path(output_dir)
    charts_dir = output_dir / 'charts'
    charts_dir.mkdir(parents=True, exist_ok=True)

    df = load_data(input_path)

    dist_df = compute_rating_distribution(df)
    summary_stats = compute_summary_stats(df)
    trend_df = compute_monthly_trend(df)
    theme_result = compute_theme_breakdown(df)
    length_df = compute_review_length_by_rating(df)
    profile_rows = compute_reviewer_profile(df)
    dow_df = compute_dow_pattern(df)

    chart_rating_distribution(dist_df, charts_dir / 'rating_distribution.png')
    if trend_df is not None and len(trend_df) > 1:
        chart_monthly_trend(trend_df, charts_dir / 'monthly_trend.png')
    if theme_result is not None:
        chart_theme_breakdown(theme_result[0], charts_dir / 'theme_breakdown.png')
    if length_df is not None:
        chart_review_length(length_df, charts_dir / 'review_length.png')
    if profile_rows:
        chart_reviewer_profile(profile_rows, charts_dir / 'reviewer_profile.png')
    if dow_df is not None:
        chart_dow_pattern(dow_df, charts_dir / 'dow_pattern.png')

    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Overview'
    merge_title(ws, 'A1:F1', 'Customer Feedback Analysis Report')
    ws.row_dimensions[1].height = 32

    r = 3
    for label, val in summary_stats:
        ws.cell(row=r, column=1, value=label).font = body_font(10, bold=True)
        ws.cell(row=r, column=2, value=val).font = body_font(10)
        ws.cell(row=r, column=1).fill = fill('FDF1F1')
        ws.cell(row=r, column=2).fill = fill('FFFFFF')
        ws.cell(row=r, column=1).border = thin_border()
        ws.cell(row=r, column=2).border = thin_border()
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='Rating Distribution').font = Font(name='Arial', size=12, bold=True, color=DARK_MAROON)
    r += 1
    next_row = write_table(ws, dist_df, r)
    embed_image(ws, charts_dir / 'rating_distribution.png', f'D{r}')
    set_col_widths(ws, {'A': 32, 'B': 16, 'C': 12})

    # ── Sheet 2: Trend Over Time ───────────────────────────────────────────
    if trend_df is not None:
        ws2 = wb.create_sheet('Trend Over Time')
        merge_title(ws2, 'A1:E1', 'Rating Trend Over Time')
        ws2.row_dimensions[1].height = 30
        write_table(ws2, trend_df, 3)
        embed_image(ws2, charts_dir / 'monthly_trend.png', 'E3')
        set_col_widths(ws2, {'A': 14, 'B': 16, 'C': 14})

        if dow_df is not None:
            r2 = 3 + len(trend_df) + 3
            ws2.cell(row=r2, column=1, value='Day-of-Week Pattern').font = Font(name='Arial', size=12, bold=True, color=DARK_MAROON)
            write_table(ws2, dow_df, r2 + 1)
            embed_image(ws2, charts_dir / 'dow_pattern.png', f'E{r2+1}')

    # ── Sheet 3: Negative Review Themes ────────────────────────────────────
    if theme_result is not None:
        freq_df, neg_count = theme_result
        ws3 = wb.create_sheet('Negative Themes')
        merge_title(ws3, 'A1:D1', f'Negative Review Themes (n={neg_count})')
        ws3.row_dimensions[1].height = 30
        write_table(ws3, freq_df, 3)
        embed_image(ws3, charts_dir / 'theme_breakdown.png', 'E3')
        set_col_widths(ws3, {'A': 32, 'B': 12, 'C': 20})

    # ── Sheet 4: Review Length & Reviewer Profile ──────────────────────────
    if length_df is not None or profile_rows:
        ws4 = wb.create_sheet('Reviewer Patterns')
        merge_title(ws4, 'A1:E1', 'Review Length & Reviewer Profile')
        ws4.row_dimensions[1].height = 30
        r4 = 3
        if length_df is not None:
            ws4.cell(row=r4, column=1, value='Review Length by Rating').font = Font(name='Arial', size=12, bold=True, color=DARK_MAROON)
            r4 += 1
            next_r = write_table(ws4, length_df, r4)
            embed_image(ws4, charts_dir / 'review_length.png', f'F{r4}')
            r4 = next_r + 2
        if profile_rows:
            profile_df = pd.DataFrame(profile_rows, columns=['Reviewer Type', 'Review Count', 'Avg Rating'])
            ws4.cell(row=r4, column=1, value='Reviewer Profile').font = Font(name='Arial', size=12, bold=True, color=DARK_MAROON)
            r4 += 1
            write_table(ws4, profile_df, r4)
            embed_image(ws4, charts_dir / 'reviewer_profile.png', f'F{r4}')
        set_col_widths(ws4, {'A': 22, 'B': 16, 'C': 18, 'D': 18})

    out_xlsx = output_dir / 'Feedback_Analysis_Report.xlsx'
    wb.save(out_xlsx)
    print(f"Analysed {len(df)} reviews.")
    print(f"Excel report: {out_xlsx}")
    print(f"Standalone charts: {charts_dir}")
    return out_xlsx


def main():
    parser = argparse.ArgumentParser(description="Statistical + visual analysis of customer feedback / review data.")
    parser.add_argument('input_file', help="Path to a .csv or .xlsx reviews export")
    parser.add_argument('output_dir', nargs='?', default='./feedback_analysis_output', help="Output folder (default: ./feedback_analysis_output)")
    args = parser.parse_args()
    run_analysis(args.input_file, args.output_dir)


if __name__ == '__main__':
    main()